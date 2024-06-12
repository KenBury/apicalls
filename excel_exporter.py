import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill
from base_classes import ExcelService

class ExcelServiceImpl(ExcelService):
    def __init__(self, output_path):
        self.output_path = output_path

    def export(self, dataframe: pd.DataFrame, export_type: str, column_headings: list, additional_data: list = None):
        if export_type == "basic":
            self.basic_export(dataframe, column_headings)
        elif export_type == "detailed":
            self.detailed_export(dataframe, column_headings, additional_data)
        else:
            raise ValueError(f"Unknown export type: {export_type}")

    def basic_export(self, dataframe: pd.DataFrame, column_headings: list):
        wb = Workbook()
        ws = wb.active

        # Write custom column headings
        for idx, heading in enumerate(column_headings, start=1):
            ws.cell(row=1, column=idx, value=heading)

        # Write the dataframe to the worksheet
        for r_idx, row in enumerate(dataframe_to_rows(dataframe, index=False, header=False), 2):  # Start from row 2 for data
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # Apply formatting (example: make the header bold and apply fill color)
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        for cell in ws[1]:  # Apply formatting to the first row (header)
            cell.font = header_font
            cell.fill = header_fill

        # Example of applying formatting to specific cells
        for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
            for cell in row:
                cell.font = Font(italic=True)
                if cell.value and pd.to_datetime(cell.value.split('\n')[0], errors='coerce') < pd.Timestamp('2000-01-01'):
                    cell.fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

        # Save the workbook
        wb.save(self.output_path)
        print(f"The data has been exported to {self.output_path}")

    def detailed_export(self, dataframe: pd.DataFrame, column_headings: list, additional_data: list):
        wb = Workbook()
        ws = wb.active

        # Write custom column headings
        for idx, heading in enumerate(column_headings, start=1):
            ws.cell(row=1, column=idx, value=heading)

        # Write the dataframe to the worksheet
        for r_idx, row in enumerate(dataframe_to_rows(dataframe, index=False, header=False), 2):  # Start from row 2 for data
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # Example of adding additional data
        for add_row in additional_data:
            ws.append(add_row)

        # Apply formatting (example: make the header bold and apply fill color)
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

        for cell in ws[1]:  # Apply formatting to the first row (header)
            cell.font = header_font
            cell.fill = header_fill

        # Example of different formatting for specific cells
        for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
            for cell in row:
                cell.font = Font(italic=True)
                if cell.value and pd.to_datetime(cell.value.split('\n')[0], errors='coerce') < pd.Timestamp('2000-01-01'):
                    cell.fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

        # Save the workbook
        wb.save(self.output_path)
        print(f"The data has been exported to {self.output_path}")
