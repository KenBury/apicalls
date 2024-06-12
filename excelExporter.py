import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill
from base_classes import Exporter

class ExcelExporter(Exporter):
    def __init__(self, output_path):
        self.output_path = output_path

    def export(self, dataframe: pd.DataFrame):
        wb = Workbook()
        ws = wb.active

        # Write the dataframe to the worksheet
        for r_idx, row in enumerate(dataframe_to_rows(dataframe, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # Apply formatting (example: make the header bold and apply fill color)
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        for cell in ws[1]:  # Apply formatting to the first row (header)
            cell.font = header_font
            cell.fill = header_fill

        # Example of applying formatting to specific cells
        for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
            for cell in row:
                cell.font = Font(italic=True)
                if cell.value and pd.to_datetime(cell.value) < pd.Timestamp('2000-01-01'):
                    cell.fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

        # Save the workbook
        wb.save(self.output_path)
        print(f"The data has been exported to {self.output_path}")
